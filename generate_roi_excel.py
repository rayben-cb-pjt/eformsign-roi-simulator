# -*- coding: utf-8 -*-
"""
eformsign ROI 시뮬레이터 - Excel 버전 생성기
웹 앱(types.ts, utils.ts)의 계산 로직을 Excel 수식으로 변환
"""

import openpyxl
from openpyxl.styles import (
    Font, PatternFill, Alignment, Border, Side, numbers
)
from openpyxl.utils import get_column_letter
from openpyxl.chart import BarChart, Reference
from copy import copy

def create_roi_excel(output_path: str):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "ROI 시뮬레이터"

    # ──────────────── 디자인 시스템 ────────────────
    # Colors
    BRAND_DARK = "1E3A5F"      # 진한 네이비
    BRAND_MID = "2563EB"       # 파란색
    BRAND_LIGHT = "DBEAFE"     # 연한 파랑 배경
    ACCENT_GREEN = "059669"    # 성과 초록
    ACCENT_GREEN_BG = "ECFDF5" # 연한 초록 배경
    ACCENT_AMBER = "D97706"    # 경고/주의 주황
    ACCENT_AMBER_BG = "FFFBEB" # 연한 주황 배경
    BG_LIGHT = "F8FAFC"        # 밝은 회색 배경
    BG_WHITE = "FFFFFF"
    BORDER_COLOR = "E2E8F0"    # 연한 보더
    TEXT_DARK = "1E293B"       # 진한 텍스트
    TEXT_MID = "475569"        # 중간 텍스트
    TEXT_LIGHT = "94A3B8"      # 연한 텍스트
    INPUT_BG = "FEF3C7"        # 입력 셀 배경 (노란빛)
    INPUT_BORDER_COLOR = "F59E0B"  # 입력 셀 보더

    # Fonts
    font_title = Font(name="맑은 고딕", size=18, bold=True, color=BG_WHITE)
    font_subtitle = Font(name="맑은 고딕", size=11, color="93C5FD")
    font_section = Font(name="맑은 고딕", size=13, bold=True, color=BRAND_DARK)
    font_label = Font(name="맑은 고딕", size=10, color=TEXT_DARK)
    font_label_bold = Font(name="맑은 고딕", size=10, bold=True, color=TEXT_DARK)
    font_value = Font(name="맑은 고딕", size=11, bold=True, color=BRAND_MID)
    font_input = Font(name="맑은 고딕", size=11, bold=True, color="B45309")
    font_result_big = Font(name="맑은 고딕", size=16, bold=True, color=ACCENT_GREEN)
    font_result = Font(name="맑은 고딕", size=12, bold=True, color=ACCENT_GREEN)
    font_note = Font(name="맑은 고딕", size=9, italic=True, color=TEXT_LIGHT)
    font_header_white = Font(name="맑은 고딕", size=10, bold=True, color=BG_WHITE)
    font_unit = Font(name="맑은 고딕", size=9, color=TEXT_LIGHT)

    # Fills
    fill_header = PatternFill(start_color=BRAND_DARK, end_color=BRAND_DARK, fill_type="solid")
    fill_brand_mid = PatternFill(start_color=BRAND_MID, end_color=BRAND_MID, fill_type="solid")
    fill_light = PatternFill(start_color=BG_LIGHT, end_color=BG_LIGHT, fill_type="solid")
    fill_white = PatternFill(start_color=BG_WHITE, end_color=BG_WHITE, fill_type="solid")
    fill_input = PatternFill(start_color=INPUT_BG, end_color=INPUT_BG, fill_type="solid")
    fill_brand_light = PatternFill(start_color=BRAND_LIGHT, end_color=BRAND_LIGHT, fill_type="solid")
    fill_green_bg = PatternFill(start_color=ACCENT_GREEN_BG, end_color=ACCENT_GREEN_BG, fill_type="solid")
    fill_amber_bg = PatternFill(start_color=ACCENT_AMBER_BG, end_color=ACCENT_AMBER_BG, fill_type="solid")

    # Borders
    thin_border = Border(
        left=Side(style="thin", color=BORDER_COLOR),
        right=Side(style="thin", color=BORDER_COLOR),
        top=Side(style="thin", color=BORDER_COLOR),
        bottom=Side(style="thin", color=BORDER_COLOR),
    )
    input_border = Border(
        left=Side(style="medium", color=INPUT_BORDER_COLOR),
        right=Side(style="medium", color=INPUT_BORDER_COLOR),
        top=Side(style="medium", color=INPUT_BORDER_COLOR),
        bottom=Side(style="medium", color=INPUT_BORDER_COLOR),
    )
    bottom_accent = Border(
        bottom=Side(style="medium", color=BRAND_MID)
    )

    # Alignment
    align_center = Alignment(horizontal="center", vertical="center", wrap_text=True)
    align_left = Alignment(horizontal="left", vertical="center", wrap_text=True)
    align_right = Alignment(horizontal="right", vertical="center")

    # Column widths
    ws.column_dimensions['A'].width = 3
    ws.column_dimensions['B'].width = 28
    ws.column_dimensions['C'].width = 18
    ws.column_dimensions['D'].width = 16
    ws.column_dimensions['E'].width = 18
    ws.column_dimensions['F'].width = 16
    ws.column_dimensions['G'].width = 3

    # Helper functions
    def merge_and_style(row, col_start, col_end, value, font, fill, alignment=align_left, height=None):
        cell = ws.cell(row=row, column=col_start, value=value)
        cell.font = font
        cell.fill = fill
        cell.alignment = alignment
        if col_start != col_end:
            ws.merge_cells(start_row=row, start_column=col_start, end_row=row, end_column=col_end)
        for c in range(col_start, col_end + 1):
            ws.cell(row=row, column=c).fill = fill
            ws.cell(row=row, column=c).border = thin_border
        if height:
            ws.row_dimensions[row].height = height

    def set_cell(row, col, value, font=font_label, fill=fill_white, alignment=align_left, number_format=None, border=thin_border):
        cell = ws.cell(row=row, column=col, value=value)
        cell.font = font
        cell.fill = fill
        cell.alignment = alignment
        cell.border = border
        if number_format:
            cell.number_format = number_format
        return cell

    def apply_row_fill(row, col_start, col_end, fill):
        for c in range(col_start, col_end + 1):
            ws.cell(row=row, column=c).fill = fill
            ws.cell(row=row, column=c).border = thin_border

    # ══════════════════════════════════════════════
    # SECTION 1: HEADER (rows 1-3)
    # ══════════════════════════════════════════════
    row = 1
    ws.row_dimensions[row].height = 50
    ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=6)
    cell = ws.cell(row=row, column=2, value="📊 eformsign ROI 시뮬레이터")
    cell.font = font_title
    cell.alignment = Alignment(horizontal="left", vertical="center")
    for c in range(2, 7):
        ws.cell(row=row, column=c).fill = fill_header

    row = 2
    ws.row_dimensions[row].height = 25
    ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=6)
    cell = ws.cell(row=row, column=2, value="전자계약 도입 시 절감 효과를 직접 계산해 보세요. 노란색 셀의 값을 자유롭게 변경할 수 있습니다.")
    cell.font = font_subtitle
    cell.alignment = Alignment(horizontal="left", vertical="center")
    for c in range(2, 7):
        ws.cell(row=row, column=c).fill = fill_header

    row = 3
    ws.row_dimensions[row].height = 8
    for c in range(2, 7):
        ws.cell(row=row, column=c).fill = fill_white

    # ══════════════════════════════════════════════
    # SECTION 2: 시뮬레이션 입력값 (rows 4~14)
    # ══════════════════════════════════════════════
    row = 5
    ws.row_dimensions[row].height = 30
    ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=6)
    cell = ws.cell(row=row, column=2, value="⚙️  시뮬레이션 입력값")
    cell.font = font_section
    cell.alignment = Alignment(horizontal="left", vertical="center")
    for c in range(2, 7):
        ws.cell(row=row, column=c).fill = fill_white
        ws.cell(row=row, column=c).border = bottom_accent

    row = 6
    ws.row_dimensions[row].height = 5
    
    # 고객 입력 파라미터 (노란색 셀)
    input_rows = []

    row = 7
    ws.row_dimensions[row].height = 28
    set_cell(row, 2, "📝 연간 계약 건수", font_label_bold, fill_light)
    set_cell(row, 3, 3000, font_input, fill_input, align_center, "#,##0", input_border)
    set_cell(row, 4, "건 / 년", font_unit, fill_light, align_left)
    ws.merge_cells(start_row=row, start_column=4, end_row=row, end_column=6)
    for c in range(4, 7):
        ws.cell(row=row, column=c).fill = fill_light
    input_rows.append(row)
    REF_CONTRACT = f"C{row}"  # C7

    row = 8
    ws.row_dimensions[row].height = 28
    set_cell(row, 2, "🤝 대면 계약 비율", font_label_bold, fill_white)
    set_cell(row, 3, 0.5, font_input, fill_input, align_center, "0%", input_border)
    set_cell(row, 4, "(나머지는 등기/우편 계약)", font_unit, fill_white, align_left)
    ws.merge_cells(start_row=row, start_column=4, end_row=row, end_column=6)
    for c in range(4, 7):
        ws.cell(row=row, column=c).fill = fill_white
    REF_FACE_RATIO = f"C{row}"  # C8

    row = 9
    ws.row_dimensions[row].height = 28
    set_cell(row, 2, "💰 eformsign 건당 요금", font_label_bold, fill_light)
    set_cell(row, 3, 220, font_input, fill_input, align_center, "#,##0", input_border)
    set_cell(row, 4, "원 / 건", font_unit, fill_light, align_left)
    ws.merge_cells(start_row=row, start_column=4, end_row=row, end_column=6)
    for c in range(4, 7):
        ws.cell(row=row, column=c).fill = fill_light
    REF_COST_PER_USE = f"C{row}"  # C9

    row = 10
    ws.row_dimensions[row].height = 28
    set_cell(row, 2, "🔧 연간 부가 옵션 비용", font_label_bold, fill_white)
    set_cell(row, 3, 0, font_input, fill_input, align_center, "#,##0", input_border)
    set_cell(row, 4, "원 / 년 (선택사항)", font_unit, fill_white, align_left)
    ws.merge_cells(start_row=row, start_column=4, end_row=row, end_column=6)
    for c in range(4, 7):
        ws.cell(row=row, column=c).fill = fill_white
    REF_OPTIONS = f"C{row}"  # C10

    # 구분선
    row = 11
    ws.row_dimensions[row].height = 5

    # ══════════════════════════════════════════════
    # SECTION 3: 산출 기준 (조정 가능한 상수)
    # ══════════════════════════════════════════════
    row = 12
    ws.row_dimensions[row].height = 30
    ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=6)
    cell = ws.cell(row=row, column=2, value="📐  산출 기준 (조정 가능)")
    cell.font = font_section
    cell.alignment = Alignment(horizontal="left", vertical="center")
    for c in range(2, 7):
        ws.cell(row=row, column=c).fill = fill_white
        ws.cell(row=row, column=c).border = bottom_accent

    row = 13
    ws.row_dimensions[row].height = 5

    # 기준 테이블 헤더
    row = 14
    ws.row_dimensions[row].height = 25
    set_cell(row, 2, "항목", font_header_white, fill_brand_mid, align_center)
    set_cell(row, 3, "기준값", font_header_white, fill_brand_mid, align_center)
    set_cell(row, 4, "단위", font_header_white, fill_brand_mid, align_center)
    set_cell(row, 5, "비고", font_header_white, fill_brand_mid, align_center)
    ws.merge_cells(start_row=row, start_column=5, end_row=row, end_column=6)
    for c in range(5, 7):
        ws.cell(row=row, column=c).fill = fill_brand_mid

    # 상수 데이터
    constants_data = [
        ("시간당 인건비", 30000, "원/시간", "사무직 평균 연봉 기준"),
        ("대면 계약 소요시간", 1.5, "시간", "이동+미팅 시간 포함"),
        ("등기 계약 소요시간", 0.33, "시간", "출력~발송~보관 (약 20분)"),
        ("전자계약 소요시간", 0.05, "시간", "문서 업로드~발송 (약 3분)"),
        ("대면 계약 실비", 15000, "원/건", "교통비, 식비 등"),
        ("등기 계약 실비", 5000, "원/건", "등기우편, 자재비 등"),
        ("계약서 종이 매수", 2, "장/건", "A4 기준"),
        ("1일 근무시간", 8, "시간", ""),
        ("연간 근무일수", 250, "일", ""),
    ]

    const_start_row = 15
    for i, (label, value, unit, note) in enumerate(constants_data):
        r = const_start_row + i
        ws.row_dimensions[r].height = 24
        bg = fill_light if i % 2 == 0 else fill_white
        set_cell(r, 2, label, font_label, bg)
        fmt = "#,##0" if isinstance(value, int) or value >= 1 else "0.00"
        set_cell(r, 3, value, font_input, fill_input, align_center, fmt, input_border)
        set_cell(r, 4, unit, font_unit, bg, align_center)
        ws.merge_cells(start_row=r, start_column=5, end_row=r, end_column=6)
        set_cell(r, 5, note, font_note, bg, align_left)
        for c in range(5, 7):
            ws.cell(row=r, column=c).fill = bg

    # Cell references for constants
    REF_HOURLY_WAGE = f"C{const_start_row}"       # C15
    REF_FACE_TIME = f"C{const_start_row + 1}"     # C16
    REF_MAIL_TIME = f"C{const_start_row + 2}"     # C17
    REF_ESIGN_TIME = f"C{const_start_row + 3}"    # C18
    REF_FACE_COST = f"C{const_start_row + 4}"     # C19
    REF_MAIL_COST = f"C{const_start_row + 5}"     # C20
    REF_A4_PER = f"C{const_start_row + 6}"        # C21
    REF_WORK_HOURS = f"C{const_start_row + 7}"    # C22
    REF_WORK_DAYS = f"C{const_start_row + 8}"     # C23

    # ESG 상수
    row_esg_start = const_start_row + len(constants_data)
    ws.row_dimensions[row_esg_start].height = 5  # spacer
    row_esg_start += 1

    esg_constants = [
        ("A4 1장당 탄소배출량", 0.00864, "kg", "환경부 탄소발자국 기준"),
        ("A4 1장당 물 사용량", 10, "L", "국제 표준 데이터"),
        ("나무 1그루당 A4 매수", 2778, "장", "1/0.00036 환산"),
    ]

    row_esg_header = row_esg_start
    ws.row_dimensions[row_esg_header].height = 25
    set_cell(row_esg_header, 2, "ESG 항목", font_header_white, fill_brand_mid, align_center)
    set_cell(row_esg_header, 3, "기준값", font_header_white, fill_brand_mid, align_center)
    set_cell(row_esg_header, 4, "단위", font_header_white, fill_brand_mid, align_center)
    set_cell(row_esg_header, 5, "비고", font_header_white, fill_brand_mid, align_center)
    ws.merge_cells(start_row=row_esg_header, start_column=5, end_row=row_esg_header, end_column=6)
    for c in range(5, 7):
        ws.cell(row=row_esg_header, column=c).fill = fill_brand_mid

    for i, (label, value, unit, note) in enumerate(esg_constants):
        r = row_esg_header + 1 + i
        ws.row_dimensions[r].height = 24
        bg = fill_light if i % 2 == 0 else fill_white
        set_cell(r, 2, label, font_label, bg)
        fmt = "0.00000" if value < 0.01 else "0.00" if value < 1 else "#,##0"
        set_cell(r, 3, value, font_input, fill_input, align_center, fmt, input_border)
        set_cell(r, 4, unit, font_unit, bg, align_center)
        ws.merge_cells(start_row=r, start_column=5, end_row=r, end_column=6)
        set_cell(r, 5, note, font_note, bg, align_left)
        for c in range(5, 7):
            ws.cell(row=r, column=c).fill = bg

    REF_CO2_PER_A4 = f"C{row_esg_header + 1}"     # CO2
    REF_WATER_PER_A4 = f"C{row_esg_header + 2}"   # Water
    REF_TREE_A4 = f"C{row_esg_header + 3}"         # Tree

    # ══════════════════════════════════════════════
    # SECTION 4: 계산 결과 (자동 계산)
    # ══════════════════════════════════════════════
    result_start = row_esg_header + 1 + len(esg_constants) + 2
    
    row = result_start
    ws.row_dimensions[row].height = 30
    ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=6)
    cell = ws.cell(row=row, column=2, value="📈  비용 분석 결과")
    cell.font = font_section
    cell.alignment = Alignment(horizontal="left", vertical="center")
    for c in range(2, 7):
        ws.cell(row=row, column=c).fill = fill_white
        ws.cell(row=row, column=c).border = bottom_accent

    row = result_start + 1
    ws.row_dimensions[row].height = 5

    # 비용 테이블 헤더
    row = result_start + 2
    ws.row_dimensions[row].height = 28
    set_cell(row, 2, "구분", font_header_white, fill_brand_mid, align_center)
    set_cell(row, 3, "도입 전 (현재)", font_header_white, fill_brand_mid, align_center)
    set_cell(row, 4, "도입 후 (eformsign)", font_header_white, fill_brand_mid, align_center)
    ws.merge_cells(start_row=row, start_column=4, end_row=row, end_column=5)
    set_cell(row, 6, "절감 효과", font_header_white, PatternFill(start_color=ACCENT_GREEN, end_color=ACCENT_GREEN, fill_type="solid"), align_center)
    for c in range(4, 6):
        ws.cell(row=row, column=c).fill = fill_brand_mid

    # 건당 비용 (도입 전)
    # costFaceBefore = FACE_OUT_OF_POCKET + (FACE_TIME * HOURLY_WAGE)
    # costMailBefore = MAIL_OUT_OF_POCKET + (MAIL_TIME * HOURLY_WAGE)
    # avgCostBefore = (costFaceBefore * faceRatio) + (costMailBefore * (1-faceRatio))
    r = result_start + 3
    ws.row_dimensions[r].height = 28
    set_cell(r, 2, "건당 비용", font_label_bold, fill_light)
    # 도입 전 건당 = (대면실비 + 대면시간*시급)*대면비율 + (등기실비 + 등기시간*시급)*(1-대면비율)
    formula_before_per = f"=({REF_FACE_COST}+{REF_FACE_TIME}*{REF_HOURLY_WAGE})*{REF_FACE_RATIO}+({REF_MAIL_COST}+{REF_MAIL_TIME}*{REF_HOURLY_WAGE})*(1-{REF_FACE_RATIO})"
    set_cell(r, 3, formula_before_per, font_value, fill_light, align_center, "#,##0 원")
    # 도입 후 건당 = 건당요금 + (옵션/건수) + 전자계약시간*시급
    formula_after_per = f"={REF_COST_PER_USE}+IF({REF_CONTRACT}>0,{REF_OPTIONS}/{REF_CONTRACT},0)+{REF_ESIGN_TIME}*{REF_HOURLY_WAGE}"
    ws.merge_cells(start_row=r, start_column=4, end_row=r, end_column=5)
    set_cell(r, 4, formula_after_per, font_value, fill_light, align_center, "#,##0 원")
    for c in range(4, 6):
        ws.cell(row=r, column=c).fill = fill_light
    # 절감 = 도입전 - 도입후
    set_cell(r, 6, f"=C{r}-D{r}", Font(name="맑은 고딕", size=11, bold=True, color=ACCENT_GREEN), fill_green_bg, align_center, "#,##0 원")

    # 연간 총비용
    r2 = r + 1
    ws.row_dimensions[r2].height = 28
    set_cell(r2, 2, "연간 총비용", font_label_bold, fill_white)
    set_cell(r2, 3, f"=C{r}*{REF_CONTRACT}", font_value, fill_white, align_center, "#,##0 원")
    ws.merge_cells(start_row=r2, start_column=4, end_row=r2, end_column=5)
    set_cell(r2, 4, f"=D{r}*{REF_CONTRACT}", font_value, fill_white, align_center, "#,##0 원")
    for c in range(4, 6):
        ws.cell(row=r2, column=c).fill = fill_white
    set_cell(r2, 6, f"=C{r2}-D{r2}", font_result, fill_green_bg, align_center, "#,##0 원")

    # ROI (%)
    r3 = r2 + 1
    ws.row_dimensions[r3].height = 28
    set_cell(r3, 2, "ROI (%)", font_label_bold, fill_light)
    ws.merge_cells(start_row=r3, start_column=3, end_row=r3, end_column=5)
    set_cell(r3, 3, f"=IF(D{r2}>0,F{r2}/D{r2}*100,0)", font_result_big, fill_light, align_center, '#,##0.0 "%"')
    for c in range(3, 6):
        ws.cell(row=r3, column=c).fill = fill_light
    # 투자 회수 기간
    set_cell(r3, 6, f'=IF(F{r2}>0,D{r2}/(F{r2}/365),0)', font_value, fill_amber_bg, align_center, '#,##0 "일"')

    # 투자회수기간 라벨
    r3b = r3 + 1
    ws.row_dimensions[r3b].height = 18
    ws.merge_cells(start_row=r3b, start_column=2, end_row=r3b, end_column=5)
    set_cell(r3b, 2, "", font_note, fill_white)
    set_cell(r3b, 6, "← 투자 회수 기간", font_note, fill_white, align_left)

    # ══════════════════════════════════════════════
    # SECTION 5: 시간 절감 분석
    # ══════════════════════════════════════════════
    time_start = r3b + 2

    row = time_start
    ws.row_dimensions[row].height = 30
    ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=6)
    cell = ws.cell(row=row, column=2, value="⏱️  시간 절감 분석")
    cell.font = font_section
    cell.alignment = Alignment(horizontal="left", vertical="center")
    for c in range(2, 7):
        ws.cell(row=row, column=c).fill = fill_white
        ws.cell(row=row, column=c).border = bottom_accent

    row = time_start + 1
    ws.row_dimensions[row].height = 5

    # 시간 테이블 헤더
    row = time_start + 2
    ws.row_dimensions[row].height = 28
    set_cell(row, 2, "구분", font_header_white, fill_brand_mid, align_center)
    set_cell(row, 3, "도입 전", font_header_white, fill_brand_mid, align_center)
    set_cell(row, 4, "도입 후", font_header_white, fill_brand_mid, align_center)
    ws.merge_cells(start_row=row, start_column=4, end_row=row, end_column=5)
    set_cell(row, 6, "절감 시간", font_header_white, PatternFill(start_color=ACCENT_GREEN, end_color=ACCENT_GREEN, fill_type="solid"), align_center)
    for c in range(4, 6):
        ws.cell(row=row, column=c).fill = fill_brand_mid

    # 건당 소요시간 (시간)
    tr = time_start + 3
    ws.row_dimensions[tr].height = 28
    set_cell(tr, 2, "건당 소요시간", font_label_bold, fill_light)
    # avgTimeBefore = FACE_TIME*faceRatio + MAIL_TIME*(1-faceRatio)
    set_cell(tr, 3, f"={REF_FACE_TIME}*{REF_FACE_RATIO}+{REF_MAIL_TIME}*(1-{REF_FACE_RATIO})", font_value, fill_light, align_center, '0.00 "시간"')
    ws.merge_cells(start_row=tr, start_column=4, end_row=tr, end_column=5)
    set_cell(tr, 4, f"={REF_ESIGN_TIME}", font_value, fill_light, align_center, '0.00 "시간"')
    for c in range(4, 6):
        ws.cell(row=tr, column=c).fill = fill_light
    set_cell(tr, 6, f"=C{tr}-D{tr}", Font(name="맑은 고딕", size=11, bold=True, color=ACCENT_GREEN), fill_green_bg, align_center, '0.00 "시간"')

    # 연간 총 소요시간 (시간)
    tr2 = tr + 1
    ws.row_dimensions[tr2].height = 28
    set_cell(tr2, 2, "연간 총 소요시간", font_label_bold, fill_white)
    set_cell(tr2, 3, f"=C{tr}*{REF_CONTRACT}", font_value, fill_white, align_center, '#,##0 "시간"')
    ws.merge_cells(start_row=tr2, start_column=4, end_row=tr2, end_column=5)
    set_cell(tr2, 4, f"=D{tr}*{REF_CONTRACT}", font_value, fill_white, align_center, '#,##0 "시간"')
    for c in range(4, 6):
        ws.cell(row=tr2, column=c).fill = fill_white
    set_cell(tr2, 6, f"=C{tr2}-D{tr2}", font_result, fill_green_bg, align_center, '#,##0 "시간"')

    # 근무일수 환산
    tr3 = tr2 + 1
    ws.row_dimensions[tr3].height = 28
    set_cell(tr3, 2, "근무일 환산", font_label_bold, fill_light)
    set_cell(tr3, 3, f"=C{tr2}/{REF_WORK_HOURS}", font_value, fill_light, align_center, '#,##0.0 "일"')
    ws.merge_cells(start_row=tr3, start_column=4, end_row=tr3, end_column=5)
    set_cell(tr3, 4, f"=D{tr2}/{REF_WORK_HOURS}", font_value, fill_light, align_center, '#,##0.0 "일"')
    for c in range(4, 6):
        ws.cell(row=tr3, column=c).fill = fill_light
    set_cell(tr3, 6, f"=C{tr3}-D{tr3}", font_result, fill_green_bg, align_center, '#,##0.0 "일"')

    # ══════════════════════════════════════════════
    # SECTION 6: ESG 탄소중립 효과
    # ══════════════════════════════════════════════
    esg_start = tr3 + 2

    row = esg_start
    ws.row_dimensions[row].height = 30
    ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=6)
    cell = ws.cell(row=row, column=2, value="🌱  ESG 탄소중립 효과")
    cell.font = font_section
    cell.alignment = Alignment(horizontal="left", vertical="center")
    for c in range(2, 7):
        ws.cell(row=row, column=c).fill = fill_white
        ws.cell(row=row, column=c).border = bottom_accent

    row = esg_start + 1
    ws.row_dimensions[row].height = 5

    # ESG 테이블 헤더
    row = esg_start + 2
    ws.row_dimensions[row].height = 28
    set_cell(row, 2, "항목", font_header_white, fill_brand_mid, align_center)
    set_cell(row, 3, "산출식", font_header_white, fill_brand_mid, align_center)
    ws.merge_cells(start_row=row, start_column=3, end_row=row, end_column=4)
    set_cell(row, 5, "절감량", font_header_white, PatternFill(start_color=ACCENT_GREEN, end_color=ACCENT_GREEN, fill_type="solid"), align_center)
    ws.merge_cells(start_row=row, start_column=5, end_row=row, end_column=6)
    for c in range(3, 5):
        ws.cell(row=row, column=c).fill = fill_brand_mid
    for c in range(5, 7):
        ws.cell(row=row, column=c).fill = PatternFill(start_color=ACCENT_GREEN, end_color=ACCENT_GREEN, fill_type="solid")

    # 종이 절감량 = 계약건수 * A4매수 + 계약건수 * (1-대면비율) * 1 (봉투)
    er = esg_start + 3
    ws.row_dimensions[er].height = 28
    set_cell(er, 2, "📄 종이 절감량", font_label_bold, fill_light)
    ws.merge_cells(start_row=er, start_column=3, end_row=er, end_column=4)
    set_cell(er, 3, f"=계약건수×A4매수 + 등기봉투", font_note, fill_light, align_center)
    for c in range(3, 5):
        ws.cell(row=er, column=c).fill = fill_light
    ws.merge_cells(start_row=er, start_column=5, end_row=er, end_column=6)
    set_cell(er, 5, f"={REF_CONTRACT}*{REF_A4_PER}+{REF_CONTRACT}*(1-{REF_FACE_RATIO})*1", font_result, fill_green_bg, align_center, '#,##0 "장"')
    for c in range(5, 7):
        ws.cell(row=er, column=c).fill = fill_green_bg

    # CO2 절감
    er2 = er + 1
    ws.row_dimensions[er2].height = 28
    set_cell(er2, 2, "🌍 CO₂ 절감량", font_label_bold, fill_white)
    ws.merge_cells(start_row=er2, start_column=3, end_row=er2, end_column=4)
    set_cell(er2, 3, "=종이절감량 × 장당 CO₂", font_note, fill_white, align_center)
    for c in range(3, 5):
        ws.cell(row=er2, column=c).fill = fill_white
    ws.merge_cells(start_row=er2, start_column=5, end_row=er2, end_column=6)
    set_cell(er2, 5, f"=E{er}*{REF_CO2_PER_A4}", font_result, fill_green_bg, align_center, '#,##0.00 "kg"')
    for c in range(5, 7):
        ws.cell(row=er2, column=c).fill = fill_green_bg

    # 물 절감
    er3 = er2 + 1
    ws.row_dimensions[er3].height = 28
    set_cell(er3, 2, "💧 물 절감량", font_label_bold, fill_light)
    ws.merge_cells(start_row=er3, start_column=3, end_row=er3, end_column=4)
    set_cell(er3, 3, "=종이절감량 × 장당 물", font_note, fill_light, align_center)
    for c in range(3, 5):
        ws.cell(row=er3, column=c).fill = fill_light
    ws.merge_cells(start_row=er3, start_column=5, end_row=er3, end_column=6)
    set_cell(er3, 5, f"=E{er}*{REF_WATER_PER_A4}", font_result, fill_green_bg, align_center, '#,##0 "L"')
    for c in range(5, 7):
        ws.cell(row=er3, column=c).fill = fill_green_bg

    # 나무 절감
    er4 = er3 + 1
    ws.row_dimensions[er4].height = 28
    set_cell(er4, 2, "🌳 나무 절감량", font_label_bold, fill_white)
    ws.merge_cells(start_row=er4, start_column=3, end_row=er4, end_column=4)
    set_cell(er4, 3, "=종이절감량 × 0.00036", font_note, fill_white, align_center)
    for c in range(3, 5):
        ws.cell(row=er4, column=c).fill = fill_white
    ws.merge_cells(start_row=er4, start_column=5, end_row=er4, end_column=6)
    set_cell(er4, 5, f"=E{er}*0.00036", font_result, fill_green_bg, align_center, '0.00 "그루"')
    for c in range(5, 7):
        ws.cell(row=er4, column=c).fill = fill_green_bg

    # ══════════════════════════════════════════════
    # SECTION 7: 핵심 요약 (KPI)
    # ══════════════════════════════════════════════
    kpi_start = er4 + 2

    row = kpi_start
    ws.row_dimensions[row].height = 30
    ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=6)
    cell = ws.cell(row=row, column=2, value="🏆  핵심 요약")
    cell.font = font_section
    cell.alignment = Alignment(horizontal="left", vertical="center")
    for c in range(2, 7):
        ws.cell(row=row, column=c).fill = fill_white
        ws.cell(row=row, column=c).border = bottom_accent

    row = kpi_start + 1
    ws.row_dimensions[row].height = 5

    # KPI 카드 형태
    kpi_items = [
        ("💰 연간 절감 비용", f"=F{r2}", "#,##0 원"),
        ("📊 투자수익률 (ROI)", f"=C{r3}/100", "0.0%"),
        ("⏱️ 절감 근무일", f"=F{tr3}", '#,##0.0 "일"'),
        ("📄 종이 절감량", f"=E{er}", '#,##0 "장"'),
        ("🌍 탄소 절감량", f"=E{er2}", '#,##0.00 "kg"'),
    ]

    for i, (label, formula, fmt) in enumerate(kpi_items):
        kr = kpi_start + 2 + i
        ws.row_dimensions[kr].height = 32
        bg = fill_light if i % 2 == 0 else fill_white
        set_cell(kr, 2, label, font_label_bold, bg)
        ws.merge_cells(start_row=kr, start_column=3, end_row=kr, end_column=6)
        set_cell(kr, 3, formula, font_result_big if i == 0 else font_result, bg, align_center, fmt)
        for c in range(3, 7):
            ws.cell(row=kr, column=c).fill = bg

    # ══════════════════════════════════════════════
    # FOOTER
    # ══════════════════════════════════════════════
    footer_row = kpi_start + 2 + len(kpi_items) + 1
    ws.row_dimensions[footer_row].height = 5

    footer_row += 1
    ws.row_dimensions[footer_row].height = 35
    ws.merge_cells(start_row=footer_row, start_column=2, end_row=footer_row, end_column=6)
    cell = ws.cell(row=footer_row, column=2, value="ⓘ 본 시뮬레이션은 입력 조건에 따른 추정치이며, 실제 도입 효과와 차이가 있을 수 있습니다.  |  © FORCS Co., Ltd.")
    cell.font = font_note
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    for c in range(2, 7):
        ws.cell(row=footer_row, column=c).fill = fill_light
        ws.cell(row=footer_row, column=c).border = thin_border

    # ══════════════════════════════════════════════
    # Print settings
    # ══════════════════════════════════════════════
    ws.sheet_properties.pageSetUpPr = openpyxl.worksheet.properties.PageSetupProperties(fitToPage=True)
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0
    ws.page_setup.orientation = "portrait"
    ws.page_margins.left = 0.5
    ws.page_margins.right = 0.5
    ws.page_margins.top = 0.5
    ws.page_margins.bottom = 0.5

    # Protect non-input cells (optional: allow yellow cells editing)
    ws.protection.sheet = False

    # Freeze top rows
    ws.freeze_panes = "B4"

    wb.save(output_path)
    print(f"✅ Excel 파일 생성 완료: {output_path}")
    return output_path


if __name__ == "__main__":
    import os
    output_dir = os.path.dirname(os.path.abspath(__file__))
    output_file = os.path.join(output_dir, "eformsign_ROI_시뮬레이터.xlsx")
    create_roi_excel(output_file)
